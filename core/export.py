#!/usr/bin/env python3
"""
数据导出功能
支持导出为 CSV、Excel 格式
"""

import csv
import os
from datetime import datetime


def export_to_csv(start_date, end_date, output_path, include_columns=None):
    """
    导出数据为 CSV 文件
    
    Args:
        start_date: 'YYYY-MM-DD HH:MM:SS'
        end_date: 'YYYY-MM-DD HH:MM:SS'
        output_path: 输出文件路径
        include_columns: 要导出的列，默认 ['timestamp', 'app_name', 'file_path', 'duration']
    
    Returns:
        dict: {'success': bool, 'row_count': int, 'file_path': str}
    """
    from core.database import query_activity_log
    
    if include_columns is None:
        include_columns = ['timestamp', 'app_name', 'file_path', 'duration']
    
    try:
        # 查询数据
        data = query_activity_log(start_date, end_date, columns=include_columns)
        
        if not data:
            return {'success': False, 'error': '没有数据可导出'}
        
        # 写入 CSV
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # 写入表头
            writer.writerow(include_columns)
            
            # 写入数据
            for row in data:
                writer.writerow(row)
        
        print(f"✅ 成功导出 {len(data)} 条记录到 {output_path}")
        return {'success': True, 'row_count': len(data), 'file_path': output_path}
        
    except Exception as e:
        print(f"❌ 导出失败：{e}")
        return {'success': False, 'error': str(e)}


def export_to_excel(start_date, end_date, output_path, include_sheets=None):
    """
    导出数据为 Excel 文件（需要 openpyxl 库）
    
    Args:
        start_date: 'YYYY-MM-DD HH:MM:SS'
        end_date: 'YYYY-MM-DD HH:MM:SS'
        output_path: 输出文件路径
        include_sheets: 包含的工作表，默认 ['raw_data', 'summary']
            - raw_data: 原始数据
            - summary: 统计摘要
    
    Returns:
        dict: {'success': bool, 'row_count': int, 'file_path': str}
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("❌ 需要安装 openpyxl 库：pip install openpyxl")
        return {'success': False, 'error': '缺少 openpyxl 库'}
    
    from core.database import query_activity_log, query_activity_stats
    
    if include_sheets is None:
        include_sheets = ['raw_data', 'summary']
    
    try:
        # 创建工作簿
        wb = Workbook()
        
        # 1. 原始数据工作表
        if 'raw_data' in include_sheets:
            ws_raw = wb.active
            ws_raw.title = "原始数据"
            
            # 查询数据
            data = query_activity_log(start_date, end_date)
            
            if data:
                # 表头
                headers = ['时间戳', '应用名称', '文件路径', '时长 (秒)']
                ws_raw.append(headers)
                
                # 设置表头样式
                header_font = Font(bold=True)
                header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
                header_alignment = Alignment(horizontal='center')
                
                for cell in ws_raw[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                
                # 数据行
                for row in data:
                    ws_raw.append(row)
                
                # 设置列宽
                for i, width in enumerate([22, 20, 40, 12], 1):
                    col_letter = get_column_letter(i)
                    ws_raw.column_dimensions[col_letter].width = width
                
                # 设置自动筛选
                ws_raw.auto_filter.ref = ws_raw.dimensions
                
                # 冻结首行
                ws_raw.freeze_panes = "A2"
        
        # 2. 统计摘要工作表
        if 'summary' in include_sheets:
            ws_summary = wb.create_sheet("统计摘要")
            
            # 按应用分组统计
            stats_by_app = query_activity_stats(start_date, end_date, group_by="app_name")
            
            # 标题
            ws_summary.merge_cells('A1:D1')
            title_cell = ws_summary['A1']
            title_cell.value = f"时间范围统计报告\n{start_date[:10]} 至 {end_date[:10]}"
            title_cell.font = Font(bold=True, size=16)
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws_summary.row_dimensions[1].height = 50
            
            ws_summary.row_dimensions[3].height = 30
            headers = ['应用名称', '总时长 (秒)', '总时长 (小时)', '记录数']
            ws_summary.append(headers)
            
            # 设置表头样式
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            for cell in ws_summary[3]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 数据行
            if stats_by_app:
                for stat in stats_by_app:
                    app_name, total_seconds, record_count = stat
                    total_hours = round(total_seconds / 3600, 2)
                    ws_summary.append([app_name, total_seconds, total_hours, record_count])
                
                # 设置列宽
                for i, width in enumerate([25, 15, 15, 12], 1):
                    col_letter = get_column_letter(i)
                    ws_summary.column_dimensions[col_letter].width = width
                
                # 添加边框
                from openpyxl.styles import Border, Side
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                
                for row in ws_summary.iter_rows(min_row=3, max_row=ws_summary.max_row, min_col=1, max_col=4):
                    for cell in row:
                        cell.border = thin_border
            
            # 添加总计
            if stats_by_app:
                total_row = ws_summary.max_row + 1
                total_seconds = sum(s[1] for s in stats_by_app)
                total_records = sum(s[2] for s in stats_by_app)
                ws_summary.append(['总计', total_seconds, round(total_seconds/3600, 2), total_records])
                
                # 设置总计行样式
                for cell in ws_summary[total_row]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        
        # 保存文件
        wb.save(output_path)
        
        row_count = len(data) if 'raw_data' in include_sheets else 0
        print(f"✅ 成功导出 Excel 文件：{output_path}")
        return {'success': True, 'row_count': row_count, 'file_path': output_path}
        
    except Exception as e:
        print(f"❌ 导出失败：{e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


def export_summary_report(start_date, end_date, output_path):
    """
    导出简要统计报告（文本格式）
    
    Args:
        start_date: 'YYYY-MM-DD HH:MM:SS'
        end_date: 'YYYY-MM-DD HH:MM:SS'
        output_path: 输出文件路径
    
    Returns:
        dict: {'success': bool, 'file_path': str}
    """
    from core.database import query_activity_stats, query_activity_log
    
    try:
        # 获取统计数据
        stats_by_app = query_activity_stats(start_date, end_date, group_by="app_name")
        data = query_activity_log(start_date, end_date)
        
        # 计算总计
        total_seconds = sum(s[1] for s in stats_by_app) if stats_by_app else 0
        total_records = len(data) if data else 0
        
        # 生成报告
        report = []
        report.append("=" * 60)
        report.append("FocusFlow 时间使用报告")
        report.append("=" * 60)
        report.append(f"\n时间范围：{start_date[:10]} 至 {end_date[:10]}")
        report.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append(f"\n总记录数：{total_records:,} 条")
        report.append(f"总时长：{total_seconds/3600:.2f} 小时 ({total_seconds/60:.1f} 分钟)")
        report.append("\n" + "-" * 60)
        report.append("应用使用时间排行：")
        report.append("-" * 60)
        report.append(f"{'应用名称':<30} {'时长 (小时)':>12} {'记录数':>10}")
        report.append("-" * 60)
        
        if stats_by_app:
            # 按时长排序
            sorted_stats = sorted(stats_by_app, key=lambda x: x[1], reverse=True)
            for app_name, total_sec, record_count in sorted_stats:
                hours = total_sec / 3600
                report.append(f"{app_name:<30} {hours:>12.2f} {record_count:>10}")
        
        report.append("\n" + "=" * 60)
        report.append("报告结束")
        report.append("=" * 60)
        
        # 写入文件
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(report))
        
        print(f"✅ 成功导出统计报告：{output_path}")
        return {'success': True, 'file_path': output_path}
        
    except Exception as e:
        print(f"❌ 导出失败：{e}")
        return {'success': False, 'error': str(e)}
